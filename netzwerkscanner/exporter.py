"""
Schreibt die Scan-Ergebnisse in die Netzwerkdoku-Vorlage (.xlsx).

Die Vorlage bleibt erhalten (Kopfzeilen, Formatierung). Die Geräte werden
ab der ersten freien Zeile unter der Spaltenüberschrift eingefügt. Felder,
die der Scanner nicht ermitteln kann, bleiben leer.
"""

from __future__ import annotations

import datetime as _dt
import os
import sys
from copy import copy
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Side

from .scanner import Host

# Spaltenreihenfolge der Vorlage (Zeile 5).
HEADER_ROW = 5
FIRST_DATA_ROW = 6

# Zuordnung: Spaltenindex (1-basiert) -> Funktion(Host) -> Zellwert
COLUMN_MAP = {
    1:  lambda h: h.ip,                                   # IP-Adresse
    2:  lambda h: h.vendor,                               # Hersteller
    3:  lambda h: h.device_type,                          # Gerätetyp
    4:  lambda h: h.hostname,                             # Netzwerkname
    5:  lambda h: h.mac,                                  # MAC-Adresse
    6:  lambda h: "",                                     # Standort (manuell)
    7:  lambda h: "",                                     # User (manuell)
    8:  lambda h: "",                                     # Kennwort (manuell)
    9:  lambda h: "",                                     # angebunden an (manuell)
    10: lambda h: h.win_function,                         # Windows Funktion
    11: lambda h: _ports_summary(h),                     # Sonstiges -> offene Ports
    12: lambda h: h.software,                             # Softwarestand
    13: lambda h: "",                                     # eingerichtet von (manuell)
}


def _ports_summary(h: Host) -> str:
    if not h.open_ports:
        return ""
    return "Offene Ports: " + ", ".join(str(p) for p in sorted(h.open_ports))


def _resource_dir() -> str:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS  # type: ignore[attr-defined]
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def default_template_path() -> str:
    return os.path.join(_resource_dir(), "data", "Netzwerkdoku-Vorlage.xlsx")


def export(hosts: List[Host], out_path: str,
           template_path: Optional[str] = None,
           kundenname: str = "", kundennummer: str = "",
           installationsdatum: str = "") -> str:
    """Füllt die Vorlage und speichert sie unter out_path. Gibt out_path zurück."""
    template_path = template_path or default_template_path()
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Kopfdaten (unter den Labels in Zeile 1 -> Zeile 2)
    if kundenname:
        ws.cell(row=2, column=1, value=kundenname)
    if kundennummer:
        ws.cell(row=2, column=2, value=kundennummer)
    ws.cell(row=2, column=3,
            value=installationsdatum or _dt.date.today().strftime("%d.%m.%Y"))

    # Randstil von der Überschrift übernehmen (für ein sauberes Tabellenbild)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = FIRST_DATA_ROW
    for h in sorted(hosts, key=lambda x: x.sort_key()):
        for col, getter in COLUMN_MAP.items():
            cell = ws.cell(row=row, column=col, value=getter(h) or None)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        row += 1

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path


def suggested_filename(kundenname: str = "") -> str:
    stamp = _dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    base = "Netzwerkdoku"
    if kundenname:
        safe = "".join(c for c in kundenname if c.isalnum() or c in " -_").strip().replace(" ", "_")
        if safe:
            base += f"_{safe}"
    return f"{base}_{stamp}.xlsx"
